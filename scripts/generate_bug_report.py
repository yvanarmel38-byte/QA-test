"""generate_bug_report.py
Reads Playwright JSON results from test-results/pw-results.json and writes
a professionally styled bug report to Documents/bug-report.xlsx.

Usage:
    pip install openpyxl
    python scripts/generate_bug_report.py
"""

import json
import re
from datetime import date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl is required. Install it with: pip install openpyxl")

BASE_DIR = Path(__file__).parent.parent
RESULTS_FILE = BASE_DIR / "test-results" / "pw-results.json"
OUTPUT_FILE = BASE_DIR / "Documents" / "bug-report.xlsx"

COLUMNS = [
    "Bug ID",
    "Test Case ID",
    "Feature Group",
    "Title / Summary",
    "Severity",
    "Priority",
    "Status",
    "Steps to Reproduce",
    "Expected Result",
    "Actual Result",
    "Browser",
    "Environment URL",
    "Date Found",
]

SEVERITY_MAP = {
    "HP_": ("High",   "P1"),
    "VC_": ("Medium", "P2"),
    "EC_": ("Low",    "P3"),
}

SEVERITY_COLOURS = {
    "High":   "FF0000",
    "Medium": "FF9900",
    "Low":    "FFFF00",
}

HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ALT_ROW_FILL = PatternFill("solid", fgColor="DCE6F1")
NORMAL_FILL  = PatternFill("solid", fgColor="FFFFFF")
THIN         = Side(style="thin")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS = {
    1: 10,   # Bug ID
    2: 14,   # TC ID
    3: 32,   # Feature Group
    4: 42,   # Title
    5: 10,   # Severity
    6: 10,   # Priority
    7: 10,   # Status
    8: 45,   # Steps
    9: 42,   # Expected
    10: 52,  # Actual
    11: 12,  # Browser
    12: 38,  # URL
    13: 13,  # Date
}


def extract_tc_id(title: str) -> str:
    match = re.search(r"((?:HP|VC|EC)_TC-\d+)", title)
    return match.group(1) if match else ""


def get_severity_priority(tc_id: str) -> tuple:
    for prefix, (sev, pri) in SEVERITY_MAP.items():
        if tc_id.startswith(prefix):
            return sev, pri
    return "Medium", "P2"


def clean_title(title: str) -> str:
    return re.sub(r"(?:HP|VC|EC)_TC-\d+\s+[–\-]\s+", "", title).strip()


ANSI_ESCAPE_RE    = re.compile(r"\x1b(?:\[[0-9;]*[mKHFJABCDsu]|[@-Z\\-_]|\[.*?[@-~])")
ILLEGAL_CHARS_RE  = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")

def sanitize(text: str) -> str:
    """Strip ANSI escape sequences then remove chars illegal in Excel cells."""
    text = ANSI_ESCAPE_RE.sub("", text or "")
    return ILLEGAL_CHARS_RE.sub("", text)


def first_meaningful_line(text: str) -> str:
    text = sanitize(text)
    skip = ("at ", "Error:", "expect(", "Received", "Expected", "Call log")
    for line in text.splitlines():
        stripped = line.strip()
        if stripped and not any(stripped.startswith(s) for s in skip):
            return stripped[:300]
    # Fallback: return the first non-empty sanitized line
    for line in text.splitlines():
        if line.strip():
            return sanitize(line.strip())[:300]
    return sanitize(text)[:300]


def collect_failures(node: dict, feature_group: str = "", results: list = None) -> list:
    if results is None:
        results = []

    title = node.get("title", "")
    if "Feature Group" in title or "Performance" in title:
        feature_group = title

    # In Playwright JSON, specs carry the test title; each spec has a tests[] array
    for spec in node.get("specs", []):
        spec_title = spec.get("title", "")
        for test in spec.get("tests", []):
            project = test.get("projectName", "")
            for result in test.get("results", []):
                if result.get("status") in ("failed", "timedOut"):
                    error_obj = result.get("error") or {}
                    raw_msg   = error_obj.get("message", "") or ""
                    results.append({
                        "title":         spec_title,
                        "feature_group": feature_group,
                        "project":       project,
                        "error":         raw_msg,  # store raw; write_report will clean
                    })

    for child in node.get("suites", []):
        collect_failures(child, feature_group, results)

    return results


def _human_actual(tc_id: str, title: str, raw_err: str) -> str:
    """Return a human-readable 'Actual Result' sentence for the bug report."""
    t = title.lower()
    err = raw_err.lower()

    # Alert explicitly captured by our helpers (submitAndExpectSuccess / expectAlertMessage)
    alert_match = re.search(
        r'validation alert:\s*["\u201c]([^"\u201d]+)["\u201d]', raw_err, re.IGNORECASE
    )
    if alert_match:
        return f'Form unexpectedly blocked submission with alert: "{alert_match.group(1).strip()}"'

    # Test expected an alert but none fired (VC tests)
    if tc_id.startswith("VC_") and "expected alert" in err:
        fragment = re.search(r"expected alert (.+)", err, re.IGNORECASE)
        reason   = fragment.group(1).strip().rstrip(".") if fragment else "for invalid input"
        return f"No validation alert was shown — form accepted the invalid input ({reason})"

    # Success-path test: p.success never appeared (form reloaded or rejected)
    if "tobevisible" in err or "p.success" in err or "locator" in err.lower():
        if "timedout" in err or "timeout" in err:
            return "Form did not display a success message within the timeout period"
        return "Form did not show a success confirmation after submitting valid data"

    # Timeout during page setup
    if "timeout" in err and "setting up" in err:
        return "Page failed to load or become interactive within the test timeout (60 s)"

    # Date-input browser sanitisation
    if "date" in t or "dd/mm" in t or "february 30" in t.lower():
        return "Browser date input accepted the invalid/non-existent date value"

    # Performance threshold exceeded
    perf_match = re.search(r"(TTFB|domContentLoaded|load)\s+(\d+)ms exceeds", raw_err, re.IGNORECASE)
    if perf_match:
        return f"Performance regression: {perf_match.group(1)} measured {perf_match.group(2)} ms, exceeding the defined threshold"

    # Generic fallback: first meaningful line
    return first_meaningful_line(raw_err) or "Test assertion failed — see Playwright HTML report"


def write_report(failures: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bug Report"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[1].height = 30

    today = date.today().strftime("%Y-%m-%d")

    for row_idx, failure in enumerate(failures, start=2):
        bug_id   = f"BUG-{row_idx - 1:03d}"
        tc_id    = extract_tc_id(failure["title"])
        severity, priority = get_severity_priority(tc_id)
        title    = clean_title(failure["title"])

        is_perf = "performance" in failure.get("feature_group", "").lower() or "timing" in title.lower()
        if is_perf:
            steps = (
                "1. Navigate to https://qa-assessment.pages.dev/\n"
                "2. Wait for the page to fully load\n"
                "3. Collect PerformanceTiming metrics from the browser"
            )
        else:
            steps = (
                f"1. Navigate to https://qa-assessment.pages.dev/\n"
                f"2. {title}\n"
                f"3. Click Submit"
            )
        if is_perf:
            expected = "All page-load timing metrics should be within defined thresholds"
        elif tc_id.startswith("VC_"):
            expected = "Form submission should be blocked with a validation alert"
        elif tc_id.startswith("EC_"):
            expected = "Edge-case input should be accepted and form should submit successfully"
        else:
            expected = "Form submits successfully and a success message is displayed"
        raw_err = sanitize(failure.get("error") or "")
        actual  = _human_actual(tc_id, title, raw_err)

        row_data = [
            bug_id,
            tc_id,
            failure["feature_group"],
            title,
            severity,
            priority,
            "New",
            steps,
            expected,
            actual,
            failure["project"],
            "https://qa-assessment.pages.dev/",
            today,
        ]

        row_fill = ALT_ROW_FILL if row_idx % 2 == 0 else NORMAL_FILL
        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=sanitize(str(value)))
            cell.fill      = row_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border    = BORDER
            if col_idx == 5 and value in SEVERITY_COLOURS:
                cell.font = Font(bold=True, color=SEVERITY_COLOURS[value], name="Calibri")
        ws.row_dimensions[row_idx].height = 75

    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nBug report written to: {OUTPUT_FILE}")
    print(f"Total bugs documented : {len(failures)}")


def main(out_path: Path = OUTPUT_FILE, extra_sources: list = None) -> None:
    global OUTPUT_FILE
    OUTPUT_FILE = out_path

    # When extra_sources are given (CI mode), merge suites from all files.
    # When running locally with no extra args, fall back to the default RESULTS_FILE.
    sources = extra_sources if extra_sources else [RESULTS_FILE]

    all_suites: list = []
    for src in sources:
        src = Path(src)
        if not src.exists():
            print(f"Warning: results file not found — skipping: {src}")
            continue
        with open(src, encoding="utf-8") as fh:
            data = json.load(fh)
        all_suites.extend(data.get("suites", []))

    if not all_suites:
        print("No test results found. Run tests first with:  npx playwright test")
        return

    failures: list = []
    for suite in all_suites:
        collect_failures(suite, "", failures)

    if not failures:
        print("All tests passed — no bugs to report.")
        print("An empty (header-only) report will be written.")

    write_report(failures)


if __name__ == "__main__":
    import sys
    args = sys.argv[1:]
    if not args:
        # Local default: single results file, default output path
        main()
    else:
        out      = Path(args[0])
        sources  = [Path(p) for p in args[1:]] if len(args) > 1 else None
        main(out_path=out, extra_sources=sources)
